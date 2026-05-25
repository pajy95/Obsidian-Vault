"""
Monte Carlo + Robustez + Análisis de Challenge FundingPips
Portfolio: NAS100 + DJI30 + XAUUSD (OOS 2025-2026)
"""
import openpyxl, warnings, random, math
warnings.filterwarnings("ignore")

# ── Extrae P&L de deals (columna 11, filas 'out') ───────────────────────────
def extract_trades(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    trades = []
    in_deals = False
    for r in range(1, ws.max_row + 1):
        r0 = str(ws.cell(r, 1).value) if ws.cell(r, 1).value else ""
        r5 = str(ws.cell(r, 5).value) if ws.cell(r, 5).value else ""
        # Detectar inicio de sección Deals (ES o EN)
        if "Transacciones" in r0 or "Deals" in r0:
            in_deals = True
            continue
        if not in_deals:
            continue
        if r5 == "out":
            pnl = ws.cell(r, 11).value
            if pnl is not None and isinstance(pnl, (int, float)):
                trades.append(float(pnl))
    return trades

# ── Archivos OOS (datos reales 2025-2026) ───────────────────────────────────
files = {
    "NAS100": r"c:\Users\JOSSE\Documents\Obsidian Vault\EA-Development\01-Strategies\BreakoutNY\Backtests\NAS100\Backtest NAS100 01_01_2025 - 28_04_2026.xlsx",
    "DJI30":  r"c:\Users\JOSSE\Documents\Obsidian Vault\EA-Development\01-Strategies\BreakoutNY\Backtests\DJI30\Backtest dji30 01_01_2025 - 28_04_2026.xlsx",
    "XAUUSD": r"c:\Users\JOSSE\Documents\Obsidian Vault\EA-Development\01-Strategies\BreakoutNY\Backtests\XAUUSD\Backtest dXAU 01_01_2025 - 28_04_2026 SET E.xlsx",
}

# ── Extraer trades individuales ──────────────────────────────────────────────
asset_trades = {}
for name, path in files.items():
    t = extract_trades(path)
    asset_trades[name] = t
    print(f"{name}: {len(t)} trades extraídos | Net=${sum(t):.2f} | WR={sum(1 for x in t if x>0)/len(t)*100:.1f}%")

# ── Portfolio combinado (todos los activos juntos, ordenados por fecha no disponible
#    → usamos bootstrap sobre cada activo por separado y combinamos por semana)
#    Aproximación conservadora: concatenar trades individuales de cada activo
all_trades = asset_trades["NAS100"] + asset_trades["DJI30"] + asset_trades["XAUUSD"]
n_total = len(all_trades)
net_total = sum(all_trades)
wr = sum(1 for x in all_trades if x > 0) / n_total
avg_win = sum(x for x in all_trades if x > 0) / max(sum(1 for x in all_trades if x > 0), 1)
avg_loss = sum(x for x in all_trades if x < 0) / max(sum(1 for x in all_trades if x < 0), 1)

print(f"\n{'='*60}")
print(f"PORTFOLIO COMBINADO OOS (NAS100 + DJI30 + XAUUSD)")
print(f"{'='*60}")
print(f"  Total trades : {n_total}")
print(f"  Net profit   : ${net_total:.2f}")
print(f"  Win Rate     : {wr*100:.1f}%")
print(f"  Avg win      : ${avg_win:.2f}")
print(f"  Avg loss     : ${avg_loss:.2f}")
print(f"  Expectancy   : ${wr*avg_win + (1-wr)*avg_loss:.2f}/trade")
pf = abs(sum(x for x in all_trades if x > 0) / sum(x for x in all_trades if x < 0))
print(f"  Profit Factor: {pf:.3f}")

# ── Monte Carlo ──────────────────────────────────────────────────────────────
random.seed(42)
N_SIM   = 10000
BALANCE = 4879.0

# Challenge FundingPips $25K: escalar risk proporcionalmente
# Cuenta $5K → $4.879, cuenta challenge $25K
# Para el análisis de la cuenta $5K actual usamos trades tal cual
# Para el challenge $25K escalamos por factor

SCALE_5K   = 1.0          # trades en cuenta $5K a escala real
SCALE_25K  = 25000 / 5000  # factor 5x para challenge $25K

TARGET_5K_PCT  = 0.0       # cuenta operativa: sin target definido (monitoreo)
TARGET_25K_F1  = 0.08      # Challenge Fase 1: +8%
TARGET_25K_F2  = 0.05      # Challenge Fase 2: +5%
DLL_PCT        = 0.05      # Daily Loss Limit 5%
MLL_PCT        = 0.10      # Max Loss Limit 10%

# Trades/mes portfolio
PERIOD_MONTHS = 16.0       # Jan 2025 – Apr 2026
TRADES_PER_MONTH = n_total / PERIOD_MONTHS

def run_monte_carlo(trades, balance_init, target_pct, mll_pct, dll_pct,
                    scale=1.0, n_sim=N_SIM, edge_decay=0.0):
    """
    edge_decay: fracción de reducción de expectancy (0.0=base, 0.30=30% peor)
    Se aplica multiplicando cada trade por (1 - edge_decay) antes de la simulación.
    """
    wins, busts, drawdowns, trades_to_target = [], [], [], []
    all_final = []

    # Aplicar decay al pool de trades
    decayed = [t * (1.0 - edge_decay) for t in trades]

    for _ in range(n_sim):
        bal = balance_init
        peak = balance_init
        max_dd = 0.0
        hit_target = False
        trades_done = 0
        daily_losses = 0.0
        trades_today = 0

        # Límite: suficientes trades para 24 meses al ritmo observado
        max_trades = int(TRADES_PER_MONTH * 24)
        shuffled = random.choices(decayed, k=max_trades)

        for pnl in shuffled:
            scaled_pnl = pnl * scale
            bal += scaled_pnl
            trades_done += 1
            trades_today += 1

            # Acumular pérdida diaria
            if scaled_pnl < 0:
                daily_losses += abs(scaled_pnl)

            # Reset diario cada ~5 trades (aprox frecuencia diaria del portfolio)
            if trades_today >= 5:
                daily_losses = 0.0
                trades_today = 0

            # Peak y DD
            if bal > peak:
                peak = bal
            dd = (peak - bal) / balance_init
            if dd > max_dd:
                max_dd = dd

            # Bust: MLL desde balance inicial o DLL desde balance_init (FundingPips: static)
            total_dd_pct = (balance_init - bal) / balance_init
            daily_dd_pct = daily_losses / balance_init
            if total_dd_pct >= mll_pct or daily_dd_pct >= dll_pct:
                busts.append(1)
                all_final.append(bal)
                break

            # Target alcanzado
            if not hit_target and target_pct > 0:
                gain = (bal - balance_init) / balance_init
                if gain >= target_pct:
                    hit_target = True
                    trades_to_target.append(trades_done)
                    all_final.append(bal)
                    break
        else:
            all_final.append(bal)

        if hit_target or (target_pct == 0):
            wins.append(bal)
        drawdowns.append(max_dd)

    n_bust = len(busts)
    n_hit  = len(trades_to_target)
    p_win  = n_hit / n_sim if target_pct > 0 else None
    p_bust = n_bust / n_sim

    avg_trades_to_target = sum(trades_to_target) / len(trades_to_target) if trades_to_target else None
    drawdowns.sort()
    p95_dd = drawdowns[int(0.95 * len(drawdowns))]
    p99_dd = drawdowns[int(0.99 * len(drawdowns))]

    return {
        "p_win": p_win,
        "p_bust": p_bust,
        "avg_trades_to_target": avg_trades_to_target,
        "p95_dd": p95_dd,
        "p99_dd": p99_dd,
        "all_final": all_final,
    }

# ── Análisis cuenta $5K operativa ───────────────────────────────────────────
print(f"\n{'='*60}")
print(f"MONTE CARLO — CUENTA OPERATIVA $5K (balance $4.879)")
print(f"{'='*60}")
res5k = run_monte_carlo(
    all_trades, BALANCE, target_pct=0.0,
    mll_pct=MLL_PCT, dll_pct=DLL_PCT, scale=SCALE_5K
)
finals = sorted(res5k["all_final"])
p10 = finals[int(0.10 * len(finals))]
p50 = finals[int(0.50 * len(finals))]
p90 = finals[int(0.90 * len(finals))]
print(f"  Probabilidad bust (MLL/DLL)  : {res5k['p_bust']*100:.1f}%")
print(f"  Max DD p95 (Monte Carlo)     : {res5k['p95_dd']*100:.2f}%")
print(f"  Max DD p99 (Monte Carlo)     : {res5k['p99_dd']*100:.2f}%")
print(f"  Balance final p10/p50/p90    : ${p10:.0f} / ${p50:.0f} / ${p90:.0f}  (en 24 meses)")

# ── Análisis Challenge $25K — 3 escenarios ───────────────────────────────────
BAL_25K = 25000.0
scenarios = [
    ("BASE (OOS histórico)",          0.00),
    ("CONSERVADOR (-20% expectancy)", 0.20),
    ("PESIMISTA (-40% expectancy)",   0.40),
]

print(f"\n{'='*60}")
print(f"MONTE CARLO — CHALLENGE $25K  (3 escenarios)")
print(f"{'='*60}")
print(f"  FundingPips: MLL=10% estático | DLL=5% | Target F1=+8% | F2=+5%")
print(f"  Escalado: risk x5 (de $5K a $25K)")

for label, decay in scenarios:
    res_f1 = run_monte_carlo(
        all_trades, BAL_25K, target_pct=TARGET_25K_F1,
        mll_pct=MLL_PCT, dll_pct=DLL_PCT, scale=SCALE_25K, edge_decay=decay
    )
    res_f2 = run_monte_carlo(
        all_trades, BAL_25K, target_pct=TARGET_25K_F1 + TARGET_25K_F2,
        mll_pct=MLL_PCT, dll_pct=DLL_PCT, scale=SCALE_25K, edge_decay=decay
    )
    months_f1 = res_f1["avg_trades_to_target"] / TRADES_PER_MONTH if res_f1["avg_trades_to_target"] else None
    months_f2 = res_f2["avg_trades_to_target"] / TRADES_PER_MONTH if res_f2["avg_trades_to_target"] else None

    print(f"\n  [{label}]")
    print(f"    P(pasar Fase 1)      : {res_f1['p_win']*100:.1f}%  |  bust: {res_f1['p_bust']*100:.1f}%")
    print(f"    P(pasar F1+F2)       : {res_f2['p_win']*100:.1f}%  |  bust: {res_f2['p_bust']*100:.1f}%")
    print(f"    Tiempo hasta Fase 1  : {months_f1:.1f} meses" if months_f1 else "    Tiempo F1: N/A")
    print(f"    Tiempo hasta Master  : {months_f2:.1f} meses" if months_f2 else "    Tiempo F1+F2: N/A")
    print(f"    Max DD p95 / p99     : {res_f2['p95_dd']*100:.2f}% / {res_f2['p99_dd']*100:.2f}%")

# ── Robustez del portfolio ───────────────────────────────────────────────────
print(f"\n{'='*60}")
print(f"ANÁLISIS DE ROBUSTEZ")
print(f"{'='*60}")

# 1. Edge Decay: primera vs segunda mitad OOS
for name, trades in asset_trades.items():
    if len(trades) < 4:
        continue
    mid = len(trades) // 2
    h1, h2 = trades[:mid], trades[mid:]
    pf1 = abs(sum(x for x in h1 if x>0) / sum(x for x in h1 if x<0)) if any(x<0 for x in h1) else 999
    pf2 = abs(sum(x for x in h2 if x>0) / sum(x for x in h2 if x<0)) if any(x<0 for x in h2) else 999
    exp1 = sum(h1)/len(h1)
    exp2 = sum(h2)/len(h2)
    decay = (pf2 - pf1) / pf1 * 100 if pf1 > 0 else 0
    print(f"\n  {name} — Edge Decay (H1 vs H2 OOS)")
    print(f"    PF H1={pf1:.3f} | PF H2={pf2:.3f} | Decay={decay:+.1f}%")
    print(f"    Exp H1=${exp1:.2f} | Exp H2=${exp2:.2f}")

# 2. Tail Risk: peores rachas consecutivas
print(f"\n  Rachas perdedoras máximas (OOS)")
for name, trades in asset_trades.items():
    max_streak = cur_streak = 0
    max_streak_usd = cur_usd = 0.0
    for t in trades:
        if t < 0:
            cur_streak += 1
            cur_usd += t
            max_streak = max(max_streak, cur_streak)
            max_streak_usd = min(max_streak_usd, cur_usd)
        else:
            cur_streak = 0
            cur_usd = 0.0
    print(f"    {name}: {max_streak} perdedores consecutivos | ${max_streak_usd:.2f}")

# 3. Correlación entre activos (simplificada)
print(f"\n  Trades por activo vs portfolio")
for name, trades in asset_trades.items():
    pct = len(trades) / n_total * 100
    net = sum(trades)
    contribution = net / net_total * 100
    print(f"    {name}: {len(trades)} trades ({pct:.0f}%) | Net=${net:.2f} ({contribution:.0f}% del profit)")

# ── Resumen ejecutivo ────────────────────────────────────────────────────────
print(f"\n{'='*60}")
print(f"RESUMEN EJECUTIVO — HONESTIDAD PROFESIONAL")
print(f"{'='*60}")
print(f"""
  Portfolio OOS (16 meses, datos reales):
  - Trades totales: {n_total} ({TRADES_PER_MONTH:.1f}/mes)
  - Expectancy: ${wr*avg_win + (1-wr)*avg_loss:.2f}/trade
  - Profit Factor combinado: {pf:.3f}

  Cuenta operativa $5K (balance $4.879):
  - Probabilidad bust: {res5k['p_bust']*100:.1f}%
  - DD máximo esperado p95: {res5k['p95_dd']*100:.2f}%

  Challenge FundingPips $25K:
  - P(pasar Fase 1): {res_f1['p_win']*100:.1f}%
  - P(pasar Fase 1+2): {res_f2['p_win']*100:.1f}%
  - Tiempo estimado hasta Master: {months_f2:.1f} meses
  - P(bust antes de pasar): {res_f2['p_bust']*100:.1f}%

  ADVERTENCIAS CRÍTICAS:
  - NAS100: 86 trades OOS (muestra válida), pero DJI30 (35) y XAU (27) son estadísticamente escasos
  - El escalado 5x para $25K amplifica tanto ganancias como pérdidas
  - Los resultados son con la frecuencia actual; en un challenge real el tiempo mínimo
    de operación puede extender el período
  - Monte Carlo asume que la distribución histórica de trades se mantiene futura
    — si el edge se degrada, las probabilidades bajan proporcionalmente
""")
