import openpyxl, warnings
warnings.filterwarnings("ignore")

def extract(path, name):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    m = {"name": name}
    for r in range(1, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, 14)]
        def v(c): return row[c] if row[c] is not None else ""
        r0 = str(row[0]) if row[0] else ""
        if "Beneficio Neto" in r0:           m["net"] = v(3)
        if "Beneficio Bruto" in r0:          m["gross"] = v(3); m["max_dd_bal"] = v(7); m["max_dd_eq"] = v(11)
        if "rdidas Brutas" in r0:            m["loss"] = v(3)
        if "Factor de Beneficio" in r0:      m["pf"] = v(3); m["exp"] = v(7)
        if "Factor de Recuperaci" in r0:     m["rec"] = v(3); m["sharpe"] = v(7)
        if "Total de operaciones" in r0:     m["trades"] = v(3)
        if "rentables (% del total)" in r0:  m["win"] = v(7); m["loss_t"] = v(11)
        if "nimo para retener" in r0:        m["hold_min"] = v(3); m["hold_avg"] = v(11)
        for c in range(4):
            val = str(row[c]) if row[c] else ""
            if "BE_BufferPct=" in val:    m["be"] = val.split("=")[1]
            if "MinSL_Points=" in val:    m["minsl"] = val.split("=")[1]
            if "MaxSL_Points=" in val:    m["maxsl"] = val.split("=")[1]
            if "EntryMaxCandle=" in val:  m["emc"] = val.split("=")[1]
            if "FilterMonday=" in val:    m["mon"] = val.split("=")[1]
            if "FilterTuesday=" in val:   m["tue"] = val.split("=")[1]
            if "FilterWednesday=" in val: m["wed"] = val.split("=")[1]
            if "FilterThursday=" in val:  m["thu"] = val.split("=")[1]
            if "FilterFriday=" in val:    m["fri"] = val.split("=")[1]
    return m

sets = {
    "SET E (solo Jue)":    r"c:\Users\JOSSE\Documents\Obsidian Vault\EA-Development\01-Strategies\BreakoutNY\Backtests\XAUUSD\Backtest dXAU 01_01_2025 - 28_04_2026 SET E.xlsx",
    "SET G (solo Jue)":    r"c:\Users\JOSSE\Documents\Obsidian Vault\EA-Development\01-Strategies\BreakoutNY\Backtests\XAUUSD\Backtest dXAU 01_01_2025 - 28_04_2026 SET G.xlsx",
    "SET E + LXJ (Lun+Mie+Jue)": r"c:\Users\JOSSE\Documents\Obsidian Vault\EA-Development\01-Strategies\BreakoutNY\Backtests\XAUUSD\Backtest dXAU 01_01_2025 - 28_04_2026 SET E + LXJ.xlsx",
}

results = {n: extract(p, n) for n, p in sets.items()}

print(f"\n{'Métrica':<22} {'SET E (Jue)':<18} {'SET G (Jue)':<18} {'SET E + LXJ':<18}")
print("-" * 76)
fields = [
    ("Días activos",  lambda m: f"L={m.get('mon','?')} X={m.get('wed','?')} J={m.get('thu','?')}"),
    ("BE_BufferPct",  lambda m: m.get('be', '')),
    ("MinSL / MaxSL", lambda m: f"{m.get('minsl','')} / {m.get('maxsl','')}"),
    ("Net Profit",    lambda m: m.get('net', '')),
    ("Profit Factor", lambda m: m.get('pf', '')),
    ("Expectancy $",  lambda m: m.get('exp', '')),
    ("Sharpe",        lambda m: m.get('sharpe', '')),
    ("Max DD bal",    lambda m: m.get('max_dd_bal', '')),
    ("Max DD eq",     lambda m: m.get('max_dd_eq', '')),
    ("Trades",        lambda m: m.get('trades', '')),
    ("Hold min/avg",  lambda m: f"{m.get('hold_min','')} / {m.get('hold_avg','')}"),
]

keys = list(results.keys())
for label, fn in fields:
    vals = [str(fn(results[k]))[:17] for k in keys]
    print(f"{label:<22} {vals[0]:<18} {vals[1]:<18} {vals[2]:<18}")
