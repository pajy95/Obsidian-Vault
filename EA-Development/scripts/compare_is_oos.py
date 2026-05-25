import openpyxl, warnings
warnings.filterwarnings("ignore")

def extract(path, label):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    m = {"label": label}
    for r in range(1, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, 14)]
        def v(c): return row[c] if row[c] is not None else ""
        r0 = str(row[0]) if row[0] else ""
        if "Beneficio Neto" in r0:           m["net"] = v(3)
        if "Beneficio Bruto" in r0:          m["gross"] = v(3); m["max_dd_bal"] = v(7); m["max_dd_eq"] = v(11)
        if "rdidas Brutas" in r0:            m["loss"] = v(3)
        if "Factor de Beneficio" in r0:      m["pf"] = round(float(v(3)), 3) if v(3) else ""; m["exp"] = v(7)
        if "Factor de Recuperaci" in r0:     m["rec"] = v(3); m["sharpe"] = v(7)
        if "Total de operaciones" in r0:     m["trades"] = v(3)
        if "rentables (% del total)" in r0:  m["win"] = v(7)
        if "nimo para retener" in r0:        m["hold_min"] = v(3); m["hold_avg"] = v(11)
        if "Per" in r0 and "odo:" in r0:     m["periodo"] = v(3)
        for c in range(4):
            val = str(row[c]) if row[c] else ""
            if "FilterMonday=" in val:    m["mon"] = val.split("=")[1]
            if "FilterWednesday=" in val: m["wed"] = val.split("=")[1]
            if "FilterThursday=" in val:  m["thu"] = val.split("=")[1]
    return m

files = {
    "Set E — IS (Jue)":      r"c:\Users\JOSSE\Documents\Obsidian Vault\EA-Development\01-Strategies\BreakoutNY\Backtests\XAUUSD\IS_XAUUSD_ThuOnly_v2.xlsx",
    "Set E — OOS (Jue)":     r"c:\Users\JOSSE\Documents\Obsidian Vault\EA-Development\01-Strategies\BreakoutNY\Backtests\XAUUSD\Backtest dXAU 01_01_2025 - 28_04_2026 SET E.xlsx",
    "Set E+LXJ — IS":        r"c:\Users\JOSSE\Documents\Obsidian Vault\EA-Development\01-Strategies\BreakoutNY\Backtests\XAUUSD\IS_XAUUSD_set_E_LXJ.xlsx",
    "Set E+LXJ — OOS":       r"c:\Users\JOSSE\Documents\Obsidian Vault\EA-Development\01-Strategies\BreakoutNY\Backtests\XAUUSD\OOS_XAUUSD_set_E_LXJ.xlsx",
}

results = {n: extract(p, n) for n, p in files.items()}

print("\n=== COMPARATIVA IS vs OOS — Set E (solo Jue) vs Set E+LXJ ===\n")
cols = list(results.keys())
print(f"{'Métrica':<22}", "  ".join(f"{c:<22}" for c in cols))
print("-" * 112)

fields = [
    ("Período",        lambda m: str(m.get("periodo", ""))[:22]),
    ("Días activos",   lambda m: f"L={m.get('mon','?')} X={m.get('wed','?')} J={m.get('thu','?')}"),
    ("Net Profit",     lambda m: str(m.get("net", ""))),
    ("Profit Factor",  lambda m: str(m.get("pf", ""))),
    ("Expectancy $",   lambda m: str(m.get("exp", ""))[:22]),
    ("Sharpe",         lambda m: str(m.get("sharpe", ""))[:22]),
    ("Max DD bal",     lambda m: str(m.get("max_dd_bal", ""))[:22]),
    ("Max DD eq",      lambda m: str(m.get("max_dd_eq", ""))[:22]),
    ("Trades",         lambda m: str(m.get("trades", ""))),
    ("Winners",        lambda m: str(m.get("win", ""))[:22]),
    ("Hold min/avg",   lambda m: f"{m.get('hold_min','')} / {m.get('hold_avg','')}"),
]

for label, fn in fields:
    vals = [fn(results[k])[:22] for k in cols]
    print(f"{label:<22}", "  ".join(f"{v:<22}" for v in vals))

# Retención OOS
print("\n--- Retención OOS (PF_OOS / PF_IS) ---")
pf_is_e   = float(results["Set E — IS (Jue)"]["pf"])   if results["Set E — IS (Jue)"]["pf"]   else 0
pf_oos_e  = float(results["Set E — OOS (Jue)"]["pf"])  if results["Set E — OOS (Jue)"]["pf"]  else 0
pf_is_lxj = float(results["Set E+LXJ — IS"]["pf"])     if results["Set E+LXJ — IS"]["pf"]     else 0
pf_oos_lxj= float(results["Set E+LXJ — OOS"]["pf"])    if results["Set E+LXJ — OOS"]["pf"]    else 0

if pf_is_e   > 0: print(f"Set E (Jue):   {pf_oos_e:.3f} / {pf_is_e:.3f} = {pf_oos_e/pf_is_e*100:.1f}%")
if pf_is_lxj > 0: print(f"Set E+LXJ:     {pf_oos_lxj:.3f} / {pf_is_lxj:.3f} = {pf_oos_lxj/pf_is_lxj*100:.1f}%")
