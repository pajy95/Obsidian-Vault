import openpyxl, warnings
warnings.filterwarnings("ignore")

def extract(path, label):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    m = {"label": label}
    for r in range(1, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, 14)]
        def v(c): return row[c] if row[c] is not None else ""
        r0 = str(row[0]) if row[0] else ""

        # Spanish labels
        if "Beneficio Neto" in r0:           m["net"] = v(3)
        if "Beneficio Bruto" in r0:          m["gross"] = v(3); m["max_dd_bal"] = v(7); m["max_dd_eq"] = v(11)
        if "rdidas Brutas" in r0:            m["loss"] = v(3)
        if "Factor de Beneficio" in r0:      m["pf"] = round(float(v(3)), 3) if v(3) else ""; m["exp"] = v(7)
        if "Factor de Recuperaci" in r0:     m["rec"] = v(3); m["sharpe"] = v(7)
        if "Total de operaciones" in r0:     m["trades"] = v(3)
        if "rentables (% del total)" in r0:  m["win"] = v(7); m["loss_t"] = v(11)
        if "nimo para retener" in r0:        m["hold_min"] = v(3); m["hold_max"] = v(7); m["hold_avg"] = v(11)
        if "Per" in r0 and "odo:" in r0:     m["periodo"] = v(3)

        # English labels
        if "Total Net Profit:" in r0:        m["net"] = v(3)
        if "Gross Profit:" in r0:            m["gross"] = v(3); m["max_dd_bal"] = v(7); m["max_dd_eq"] = v(11)
        if "Gross Loss:" in r0:              m["loss"] = v(3)
        if "Profit Factor:" in r0:
            raw = v(3)
            m["pf"] = round(float(raw), 3) if raw != "" else ""
            m["exp"] = v(7)
        if "Recovery Factor:" in r0:         m["rec"] = v(3); m["sharpe"] = v(7)
        if "Total Trades:" in r0:            m["trades"] = v(3)
        if "Profit Trades (% of total)" in r0: m["win"] = v(7); m["loss_t"] = v(11)
        if "Minimal position holding" in r0: m["hold_min"] = v(3); m["hold_max"] = v(7); m["hold_avg"] = v(11)
        if "Period:" in r0:                  m["periodo"] = v(3)

        for c in range(4):
            val = str(row[c]) if row[c] else ""
            if "BE_BufferPct=" in val:       m["be"] = val.split("=")[1]
            if "MinSL_Points=" in val:       m["minsl"] = val.split("=")[1]
            if "MaxSL_Points=" in val:       m["maxsl"] = val.split("=")[1]
            if "ATR_FilterEnable=" in val:   m["atr"] = val.split("=")[1]
            if "ATR_MaxMultiplier=" in val:  m["atr_mult"] = val.split("=")[1]
            if "FilterMonday=" in val:       m["mon"] = val.split("=")[1]
            if "FilterWednesday=" in val:    m["wed"] = val.split("=")[1]
            if "FilterFriday=" in val:       m["fri"] = val.split("=")[1]
    return m

files = {
    "IS v1 (L+M+X+V)":   r"c:\Users\JOSSE\Documents\Obsidian Vault\EA-Development\01-Strategies\BreakoutNY\Backtests\SP500\IS_SP500_MonTueWedFri_v1.xlsx",
    "OOS v1 (L+M+X+V)":  r"c:\Users\JOSSE\Documents\Obsidian Vault\EA-Development\01-Strategies\BreakoutNY\Backtests\SP500\OOS_SP500_MonTueWedFri_v1.xlsx",
    "IS v2 (L+M+X)":      r"c:\Users\JOSSE\Documents\Obsidian Vault\EA-Development\01-Strategies\BreakoutNY\Backtests\SP500\IS_SP500_MonTueWed_v2.xlsx",
    "OOS v2 (L+M+X)":     r"c:\Users\JOSSE\Documents\Obsidian Vault\EA-Development\01-Strategies\BreakoutNY\Backtests\SP500\OOS_SP500_MonTueWed_v2.xlsx",
    "IS v2 Abril":         r"c:\Users\JOSSE\Documents\Obsidian Vault\EA-Development\01-Strategies\BreakoutNY\Backtests\SP500\IS_SP500_MonTueWed_v2_abril.xlsx",
    "OOS v2 Abril":        r"c:\Users\JOSSE\Documents\Obsidian Vault\EA-Development\01-Strategies\BreakoutNY\Backtests\SP500\OOS_SP500_MonTueWed_v2_Abril.xlsx",
}

results = {n: extract(p, n) for n, p in files.items()}

print("\n=== SP500 BreakoutNY — Comparativa completa ===\n")
fields = [
    ("Periodo",         lambda m: str(m.get("periodo", ""))[:30]),
    ("Dias activos",    lambda m: f"L={m.get('mon','?')} X={m.get('wed','?')} V={m.get('fri','?')}"),
    ("BE/MinSL/MaxSL",  lambda m: f"{m.get('be','')} / {m.get('minsl','')} / {m.get('maxsl','')}"),
    ("ATR filter",      lambda m: f"{m.get('atr','')} mult={m.get('atr_mult','')}"),
    ("Net Profit",      lambda m: str(m.get("net", ""))),
    ("Profit Factor",   lambda m: str(m.get("pf", ""))),
    ("Expectancy $",    lambda m: str(m.get("exp", ""))[:20]),
    ("Sharpe",          lambda m: str(m.get("sharpe", ""))[:20]),
    ("Max DD bal",      lambda m: str(m.get("max_dd_bal", ""))[:20]),
    ("Max DD eq",       lambda m: str(m.get("max_dd_eq", ""))[:20]),
    ("Trades",          lambda m: str(m.get("trades", ""))),
    ("Winners",         lambda m: str(m.get("win", ""))[:20]),
    ("Hold max",        lambda m: str(m.get("hold_max", ""))),
    ("Hold avg",        lambda m: str(m.get("hold_avg", ""))),
]

cols = list(results.keys())
header = f"{'Metrica':<22}" + "".join(f"{c:<22}" for c in cols)
print(header)
print("-" * (22 + 22 * len(cols)))
for label, fn in fields:
    vals = [fn(results[k])[:21] for k in cols]
    print(f"{label:<22}" + "".join(f"{v:<22}" for v in vals))

# Retencion OOS
print("\n--- Retencion OOS (PF_OOS / PF_IS) ---")
pairs = [
    ("v1 L+M+X+V", "IS v1 (L+M+X+V)", "OOS v1 (L+M+X+V)"),
    ("v2 L+M+X",   "IS v2 (L+M+X)",   "OOS v2 (L+M+X)"),
    ("v2 Abril",   "IS v2 Abril",      "OOS v2 Abril"),
]
for name, is_k, oos_k in pairs:
    pf_is  = results[is_k].get("pf", 0)
    pf_oos = results[oos_k].get("pf", 0)
    if pf_is and pf_oos:
        ret = float(pf_oos) / float(pf_is) * 100
        print(f"  {name}: {pf_oos} / {pf_is} = {ret:.1f}%")
